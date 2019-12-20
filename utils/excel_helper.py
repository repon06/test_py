#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

from utils.registry import registry


class Excel():

    def read_file(self, path, debug=False):
        wb = load_workbook(path)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        row_count = sheet.max_row
        column_count = sheet.max_column
        if debug:
            print(f'строк: {row_count}, столбцов: {column_count}')
        registry_list = []
        for row in range(row_count + 1):
            if row > 1:
                if 'КД' in sheet.cell(1, 1).value and 'РПО' in sheet.cell(1, 4).value:
                    registry_list.append(registry(fio=sheet.cell(row, 2).value,
                                                  bd=str(sheet.cell(row, 3).value),
                                                  kd=sheet.cell(row, 1).value,
                                                  rpo=sheet.cell(row, 4).value))
                else:
                    print(f'структура {path} файла не соответствует реестру')
        if debug:
            print(registry_list)
        return registry_list

    def save_file(self, path, data):
        print(path)
        return True
